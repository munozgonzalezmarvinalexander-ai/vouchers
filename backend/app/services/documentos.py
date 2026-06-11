"""
services/documentos.py
----------------------
Genera las tres salidas de un voucher a partir de UNA sola plantilla:

  - HTML  : para imprimir directo en el navegador (sin descargar nada) y como
            vista previa.
  - PDF   : el mismo HTML convertido con WeasyPrint (corre en el servidor).
  - Excel : reconstruye el voucher en .xlsx (openpyxl).

WeasyPrint necesita librerías del sistema (pango/cairo) que se instalan en el
servidor Linux donde se despliega; por eso se importa de forma perezosa: si no
está disponible, la app sigue funcionando y solo el endpoint de PDF avisa.
"""

from io import BytesIO
from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

TEMPLATES = Path(__file__).resolve().parent.parent.parent / "templates"
_env = Environment(
    loader=FileSystemLoader(str(TEMPLATES)),
    autoescape=select_autoescape(["html"]),
)

ORG = "CONACMI"


# Dónde empieza el voucher desde el borde superior de la hoja carta.
# El comprobante del banco va arriba; el voucher cae en la parte de abajo.
TOPS = {"medio": "130mm", "tercio": "168mm"}


def _money(valor) -> str:
    """Formatea un monto; cadena vacía si es cero (para no ensuciar el voucher)."""
    n = float(valor or 0)
    return f"{n:,.2f}" if n else ""


def _contexto(voucher, auto_print=False, formato="medio") -> dict:
    lineas = [
        {
            "codigo": d.cuenta.codigo if d.cuenta else "",
            "descripcion": d.descripcion or (d.cuenta.nombre if d.cuenta else ""),
            "debe": _money(d.debe),
            "haber": _money(d.haber),
            "es_banco": bool(d.cuenta.es_banco) if d.cuenta else False,
        }
        for d in voucher.detalles
    ]
    return {
        "numero": voucher.numero,  # solo para el <title> de la pestaña
        "concepto": voucher.concepto,
        "lineas": lineas,
        "total_debe": _money(voucher.total_debe),
        "total_haber": _money(voucher.total_haber),
        "elaborado": voucher.elaborado_por.nombre if voucher.elaborado_por else "",
        "revisado": voucher.revisado_por.nombre if voucher.revisado_por else "",
        "autorizado": voucher.autorizado_por.nombre if voucher.autorizado_por else "",
        "auto_print": auto_print,
        "top": TOPS.get(formato, TOPS["medio"]),
    }


def render_html(voucher, auto_print: bool = False, formato: str = "medio") -> str:
    return _env.get_template("voucher.html").render(**_contexto(voucher, auto_print, formato))


def render_pdf(voucher, formato: str = "medio") -> bytes:
    html = render_html(voucher, auto_print=False, formato=formato)
    try:
        from weasyprint import HTML  # import perezoso (ver nota arriba)
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "La generación de PDF no está disponible en este servidor "
            "(falta WeasyPrint o sus librerías). La impresión directa y el Excel sí funcionan."
        ) from e
    return HTML(string=html).write_pdf()


def generar_excel(voucher) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Voucher"

    negrita = Font(bold=True)
    der = Alignment(horizontal="right")
    fmt = "#,##0.00"

    ws["A1"] = ORG
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Voucher contable · {voucher.proyecto.nombre if voucher.proyecto else ''}"

    ws["A4"] = "No.:";       ws["B4"] = voucher.numero
    ws["A5"] = "Fecha:";     ws["B5"] = str(voucher.fecha)
    ws["A6"] = "Estado:";    ws["B6"] = voucher.estado.lower()
    for fila in ("A4", "A5", "A6"):
        ws[fila].font = negrita

    ws["A8"] = "Concepto:"; ws["A8"].font = negrita
    ws["A9"] = voucher.concepto

    encabezado = 11
    for col, titulo in zip("ABCD", ("Código", "Descripción", "Debe", "Haber")):
        c = ws[f"{col}{encabezado}"]
        c.value = titulo
        c.font = negrita

    fila = encabezado + 1
    for d in voucher.detalles:
        ws[f"A{fila}"] = d.cuenta.codigo if d.cuenta else ""
        ws[f"B{fila}"] = d.descripcion or (d.cuenta.nombre if d.cuenta else "")
        debe = float(d.debe or 0)
        haber = float(d.haber or 0)
        if debe:
            ws[f"C{fila}"] = debe; ws[f"C{fila}"].number_format = fmt
        if haber:
            ws[f"D{fila}"] = haber; ws[f"D{fila}"].number_format = fmt
        fila += 1

    ws[f"B{fila}"] = "Totales"; ws[f"B{fila}"].font = negrita; ws[f"B{fila}"].alignment = der
    ws[f"C{fila}"] = float(voucher.total_debe); ws[f"C{fila}"].number_format = fmt; ws[f"C{fila}"].font = negrita
    ws[f"D{fila}"] = float(voucher.total_haber); ws[f"D{fila}"].number_format = fmt; ws[f"D{fila}"].font = negrita

    firmas = fila + 3
    ws[f"A{firmas}"] = "Elaborado por"; ws[f"B{firmas}"] = "Revisado por"; ws[f"C{firmas}"] = "Autorizado por"
    ws[f"A{firmas+1}"] = voucher.elaborado_por.nombre if voucher.elaborado_por else ""
    ws[f"B{firmas+1}"] = voucher.revisado_por.nombre if voucher.revisado_por else ""
    ws[f"C{firmas+1}"] = voucher.autorizado_por.nombre if voucher.autorizado_por else ""

    for col, ancho in {"A": 16, "B": 48, "C": 14, "D": 14}.items():
        ws.column_dimensions[col].width = ancho
    ws.page_setup.orientation = "portrait"
    ws.print_area = f"A1:D{firmas+1}"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
