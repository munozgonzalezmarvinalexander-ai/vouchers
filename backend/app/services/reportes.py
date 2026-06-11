"""
services/reportes.py
--------------------
Reportes agregados:
  - libro_vouchers     : la lista de vouchers de un proyecto/periodo y su total.
  - totales_por_partida: cuánto se gastó en cada partida (suma del debe).

Ambos ignoran los vouchers ANULADOS.
"""

from io import BytesIO

from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font

from app.models import Cuenta, Voucher, VoucherDetalle


def _filtrar(consulta, proyecto_id, desde, hasta):
    consulta = consulta.filter(Voucher.estado != "ANULADO")
    if proyecto_id:
        consulta = consulta.filter(Voucher.proyecto_id == proyecto_id)
    if desde:
        consulta = consulta.filter(Voucher.fecha >= desde)
    if hasta:
        consulta = consulta.filter(Voucher.fecha <= hasta)
    return consulta


def libro_vouchers(db, proyecto_id=None, desde=None, hasta=None):
    consulta = _filtrar(db.query(Voucher), proyecto_id, desde, hasta)
    vouchers = consulta.order_by(Voucher.fecha, Voucher.id).all()
    total = sum((v.total or 0) for v in vouchers)
    return vouchers, total


def totales_por_partida(db, proyecto_id=None, desde=None, hasta=None):
    suma = func.sum(VoucherDetalle.debe)
    consulta = (
        db.query(
            Cuenta.codigo,
            Cuenta.nombre,
            suma.label("total"),
            func.count(VoucherDetalle.id).label("veces"),
        )
        .join(VoucherDetalle, VoucherDetalle.cuenta_id == Cuenta.id)
        .join(Voucher, Voucher.id == VoucherDetalle.voucher_id)
        .filter(VoucherDetalle.debe > 0)
    )
    consulta = _filtrar(consulta, proyecto_id, desde, hasta)
    filas = consulta.group_by(Cuenta.id).order_by(suma.desc()).all()
    return [
        {"codigo": c, "nombre": n, "total": t or 0, "veces": v}
        for c, n, t, v in filas
    ]


def libro_excel(vouchers, total) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro de vouchers"
    negrita = Font(bold=True)
    fmt = "#,##0.00"

    encabezados = ["Número", "Fecha", "Concepto", "Total", "Estado"]
    for i, t in enumerate(encabezados, start=1):
        c = ws.cell(row=1, column=i, value=t)
        c.font = negrita

    fila = 2
    for v in vouchers:
        ws.cell(row=fila, column=1, value=v.numero)
        ws.cell(row=fila, column=2, value=str(v.fecha))
        ws.cell(row=fila, column=3, value=v.concepto)
        celda = ws.cell(row=fila, column=4, value=float(v.total or 0))
        celda.number_format = fmt
        ws.cell(row=fila, column=5, value=v.estado.lower())
        fila += 1

    ws.cell(row=fila + 1, column=3, value="Total").font = negrita
    tcelda = ws.cell(row=fila + 1, column=4, value=float(total))
    tcelda.number_format = fmt
    tcelda.font = negrita

    for col, ancho in {"A": 22, "B": 12, "C": 55, "D": 14, "E": 12}.items():
        ws.column_dimensions[col].width = ancho

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
