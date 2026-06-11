"""
extraer_nomenclatura.py
-----------------------
Lee la nomenclatura contable oficial (plan de cuentas jerárquico) y produce el
catálogo CANÓNICO de cuentas de detalle, listo para la base de datos.

Reglas:
  - Usa las cuentas de DETALLE (columna C, código de 8 dígitos). Las filas de
    agrupación (columna B) solo sirven para entender la jerarquía.
  - Corrige tildes de forma SEGURA: si una palabra aparece acentuada en algún
    lugar del archivo, se acentúa en TODAS sus apariciones. No inventa tildes
    de palabras que nunca aparecen acentuadas (salvo un suplemento mínimo y
    seguro).
  - Respeta los nombres tal cual: NO completa los nombres truncados (p. ej.
    "...de pl"), no cambia la redacción ni las mayúsculas/minúsculas.
  - Deduce tipo y naturaleza por la clase del código; marca los bancos
    (grupo 11102) y les extrae el número de cuenta.

Uso:
  python scripts/extraer_nomenclatura.py --input nomenclatura.xlsx --out data/catalogo_cuentas.csv
"""

import argparse
import csv
import re
import unicodedata
from collections import Counter, defaultdict

from openpyxl import load_workbook

# Clase mayor (primer dígito del código) -> (tipo, naturaleza)
CLASES = {
    "1": ("ACTIVO", "DEUDORA"),
    "2": ("PASIVO", "ACREEDORA"),
    "3": ("PATRIMONIO", "ACREEDORA"),
    "5": ("INGRESOS", "ACREEDORA"),
    "6": ("GASTOS", "DEUDORA"),
    "7": ("GASTOS", "DEUDORA"),
}

# Tildes seguras aunque la palabra nunca aparezca acentuada en el archivo.
SUPLEMENTO = {
    "telefonicas": "telefónicas", "telefonico": "telefónico", "telefonica": "telefónica",
    "numero": "número", "analisis": "análisis",
}

RE_NUM_BANCO = re.compile(r"\d[\d-]{6,}\d")     # ej. 002-005041-9, 46-00855562-00001-1
RE_PALABRA = re.compile(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]+")
RE_CODIGO = re.compile(r"\d{3,}")


def sin_tilde(texto: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", texto.lower()) if not unicodedata.combining(c))


def tiene_tilde(texto: str) -> bool:
    return any(c in "áéíóúÁÉÍÓÚ" for c in texto)


def construir_mapa_tildes(nombres) -> dict:
    """Para cada palabra, si aparece acentuada en algún nombre, esa forma manda."""
    formas = defaultdict(Counter)
    for n in nombres:
        for tok in RE_PALABRA.findall(n):
            if len(tok) > 3:
                formas[sin_tilde(tok)][tok] += 1
    mapa = dict(SUPLEMENTO)
    for base, cnt in formas.items():
        acentuadas = {f.lower(): c for f, c in cnt.items() if tiene_tilde(f)}
        if acentuadas:
            mapa[base] = Counter(acentuadas).most_common(1)[0][0]
    return mapa


def _aplicar_case(original: str, corregido: str) -> str:
    if original.isupper():
        return corregido.upper()
    if original[:1].isupper():
        return corregido[:1].upper() + corregido[1:]
    return corregido


def corregir_tildes(nombre: str, mapa: dict) -> str:
    def repl(m):
        tok = m.group(0)
        base = sin_tilde(tok)
        if base in mapa and mapa[base] != base:
            return _aplicar_case(tok, mapa[base])
        return tok
    return RE_PALABRA.sub(repl, nombre)


def clasificar(codigo: str):
    tipo, naturaleza = CLASES.get(codigo[0], ("OTRO", "DEUDORA"))
    es_banco = codigo.startswith("11102")
    if es_banco:
        tipo = "BANCO"
    return tipo, naturaleza, es_banco


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="nomenclatura.xlsx")
    ap.add_argument("--out", default="data/catalogo_cuentas.csv")
    args = ap.parse_args()

    wb = load_workbook(args.input, data_only=True)
    ws = wb[wb.sheetnames[0]]

    detalle = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        c = ws.cell(r, 3).value
        if a in (None, "") or c in (None, ""):
            continue
        codigo = str(c).strip()
        if not RE_CODIGO.fullmatch(codigo):
            continue
        detalle.append((codigo, str(a).strip()))

    mapa = construir_mapa_tildes([n for _, n in detalle])

    filas = []
    corregidos = 0
    for codigo, nombre in detalle:
        nuevo = corregir_tildes(nombre, mapa)
        if nuevo != nombre:
            corregidos += 1
        tipo, naturaleza, es_banco = clasificar(codigo)
        numero = ""
        if es_banco:
            m = RE_NUM_BANCO.search(nuevo)
            numero = m.group(0) if m else ""
        filas.append({
            "codigo": codigo,
            "nombre": nuevo,
            "tipo": tipo,
            "naturaleza": naturaleza,
            "es_banco": "1" if es_banco else "0",
            "numero_cuenta": numero,
            "resultado": "",
            "activo": "1",
        })

    filas.sort(key=lambda f: f["codigo"])
    cols = ["codigo", "nombre", "tipo", "naturaleza", "es_banco", "numero_cuenta", "resultado", "activo"]
    with open(args.out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(filas)

    print(f"Cuentas de detalle escritas  : {len(filas)}")
    print(f"Nombres con tildes corregidas: {corregidos}")
    print("Por tipo:", dict(Counter(f["tipo"] for f in filas)))
    print(f"Bancos con número extraído   : {sum(1 for f in filas if f['es_banco']=='1' and f['numero_cuenta'])}")


if __name__ == "__main__":
    main()
